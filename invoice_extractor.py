#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
发票信息提取脚本
功能：扫描目录中的所有PDF发票，提取发票号码、开票日期、购买方、销售方、金额等信息
输出：生成Excel格式的发票清单
"""

import fitz  # PyMuPDF
import os
import re
import time
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ---- 预编译正则表达式（避免每次调用重复编译）----
RE_INVOICE_20DIGIT = re.compile(r'\b(\d{20})\b')
RE_INVOICE_CODE_12 = re.compile(r'\b(\d{12})\b')
RE_INVOICE_NUMBER_8 = re.compile(r'\b(\d{8})\b')
RE_VERIFY_SPLIT = re.compile(r'(\d{5})\s+(\d{5})\s+(\d{5})\s+(\d{5})')
RE_DATE = re.compile(r'(\d{4})年(\d{1,2})月(\d{1,2})日')
RE_TAX_NUMBER = re.compile(r'\b[0-9A-Z]{18}\b')
RE_ITEM_FULL = re.compile(r'\*[^*]+\*[^\n*]+')
RE_ITEM_SHORT = re.compile(r'\*([^*]+)\*')
RE_YUANZHENG_AMOUNT = re.compile(r'圆整\s*[¥￥]?\s*([\d,]+\.?\d*)')
RE_ALL_AMOUNTS = re.compile(r'[¥￥]\s*([\d,]+\.?\d*)')
RE_FILENAME_AMOUNT = re.compile(r'(\d+\.?\d*)\.(?:PDF|PNG|JPG|JPEG)', re.IGNORECASE)
RE_HEJI_AMOUNT = re.compile(r'(?:价税合计|合计|总计|大写|小写).*?[¥￥]\s*([\d,]+\.?\d*)', re.DOTALL)

# 排除模式（预编译）
EXCLUDE_PATTERNS = [
    re.compile(p) for p in [
        r'\*[^*]+\*',
        r'项目|规格|单位|数量|单价|金额|税率|税额|合计|备注|开票人|下载次数|发票号码|开票日期',
        r'国家税务总局|发票监制|电子发票|普通发票|广东省税务局',
        r'价税合计|大写|小写',
    ]
]


def extract_invoice_info(pdf_path, buyer_keyword=None):
    """从PDF发票中提取信息

    Args:
        pdf_path: PDF文件路径
        buyer_keyword: 购买方公司名称关键词（用于识别购买方）
    """
    if buyer_keyword is None:
        buyer_keyword = ""
    try:
        doc = fitz.open(pdf_path)
        text = ""
        for page in doc:
            text += page.get_text()
        doc.close()

        info = {
            '发票号码': '',
            '发票代码': '',
            '开票日期': '',
            '购买方': '',
            '购买方税号': '',
            '销售方': '',
            '销售方税号': '',
            '项目内容': '',
            '金额': '',
            '备注': ''
        }

        # 提取发票号码 - 支持两种格式:
        # 1. 普通发票: 20位连续数字 (发票代码+发票号码的组合)
        # 2. 高速费发票: 8位发票号码 + 12位发票代码
        fp_match_20 = RE_INVOICE_20DIGIT.search(text)
        if fp_match_20:
            info['发票号码'] = fp_match_20.group(1)
        else:
            # 尝试提取高速费发票的发票代码(12位)和发票号码(8位)
            # 发票代码通常是12位数字，发票号码是8位数字
            code_match = RE_INVOICE_CODE_12.search(text)
            number_match = RE_INVOICE_NUMBER_8.search(text)
            if code_match and number_match:
                info['发票代码'] = code_match.group(1)
                info['发票号码'] = number_match.group(1)
            # 如果校验码被空格分隔，组合成20位
            verify_match = RE_VERIFY_SPLIT.search(text)
            if verify_match:
                info['校验码'] = verify_match.group(1) + verify_match.group(2) + verify_match.group(3) + verify_match.group(4)

        # 提取开票日期
        date_match = RE_DATE.search(text)
        if date_match:
            info['开票日期'] = f"{date_match.group(1)}-{date_match.group(2).zfill(2)}-{date_match.group(3).zfill(2)}"

        # 提取税号（18位，可能包含字母）
        tax_numbers = RE_TAX_NUMBER.findall(text)
        valid_taxes = [t for t in tax_numbers if not (t.isdigit() and len(t) == 20)]
        if len(valid_taxes) >= 1:
            info['购买方税号'] = valid_taxes[0]
        if len(valid_taxes) >= 2:
            info['销售方税号'] = valid_taxes[1]

        # 提取项目内容 - 支持 *类别*项目名 格式（如 *经营租赁*通行费）
        # 先尝试匹配完整的 *类别*项目名 格式
        item_match_full = RE_ITEM_FULL.search(text)
        if item_match_full:
            item_content = item_match_full.group(0).strip()
            # 去除可能的换行符和多余空格
            item_content = item_content.replace('\n', '').replace('\r', '')
            info['项目内容'] = item_content[:50]
        else:
            # 降级到只匹配 *类别* 格式
            item_match = RE_ITEM_SHORT.search(text)
            if item_match:
                info['项目内容'] = item_match.group(0)[:30]

        # 销售方关键词（扩展覆盖各种类型）
        seller_keywords = [
            '有限公司', '股份有限公司', '有限责任公司',
            '科技', '网络', '文化', '婴童',
            '贸易', '电子商务',
            '酒店', '饭店', '餐饮', '娱乐', '百货',
            '加油站', '石油化工', '石化',
            '商行',
            '实业', '制造', '加工',
            '中心', '工作室', '经营部',
            # 个体工商户常见后缀
            '工艺品厂', '木制品厂', '制品厂', '加工厂',
            '电子商务商行', '电子商务网店', '纸塑制品网店',
            '茶业有限公司', '餐饮有限公司',
        ]

        # 提取所有可能的销售方名称
        all_sellers = []
        lines = text.split('\n')
        for line in lines:
            line = line.strip()
            if len(line) < 5 or len(line) > 80:
                continue
            if any(p.search(line) for p in EXCLUDE_PATTERNS):
                continue
            if any(kw in line for kw in seller_keywords):
                if not line.endswith('费'):
                    if line not in all_sellers:
                        all_sellers.append(line)

        # 识别购买方和销售方
        # 策略：如果只有一家公司，根据购买方关键词判断
        if len(all_sellers) == 1:
            company = all_sellers[0]
            # 如果包含购买方关键词，则是购买方，销售方为空
            if buyer_keyword and buyer_keyword in company:
                info['购买方'] = company
            # 否则这家公司是销售方
            else:
                info['销售方'] = company
                # 如果购买方关键词没有匹配，使用默认值
                if buyer_keyword:
                    info['购买方'] = buyer_keyword + '（未在发票中显示）'
        elif len(all_sellers) >= 2:
            # 有多家公司时，根据购买方关键词匹配
            buyer_found = False
            if buyer_keyword:
                for c in all_sellers:
                    if buyer_keyword in c:
                        info['购买方'] = c
                        buyer_found = True
                        break
            # 销售方是第一个不同于购买方的商家
            for c in all_sellers:
                if c != info['购买方']:
                    info['销售方'] = c
                    break
            # 如果没有找到购买方，使用默认值
            if not buyer_found and buyer_keyword:
                info['购买方'] = buyer_keyword + '（未在发票中显示）'

        # 如果没找到，尝试从税号附近提取
        if not info['销售方'] and info.get('销售方税号'):
            tax = info['销售方税号']
            idx = text.find(tax)
            if idx > 0:
                context = text[max(0, idx-100):min(len(text), idx+100)]
                for kw in ['店', '商行', '有限公司', '商贸', '科技', '贸易', '酒店', '饭店', '餐饮']:
                    match = re.search(r'([^\s\n]+{kw}[^\s\n]*)'.format(kw=kw), context)
                    if match:
                        seller = match.group(1).strip('*,、。.\n\t\r')
                        if seller != info['购买方'] and len(seller) > 4:
                            info['销售方'] = seller
                            break

        # 提取金额 - 优先找"圆整"后的金额
        amount_near_yuanzheng = RE_YUANZHENG_AMOUNT.search(text)
        if amount_near_yuanzheng:
            info['金额'] = amount_near_yuanzheng.group(1).replace(',', '')
        else:
            # 第二优先：找"价税合计""合计""总计"附近的金额
            heji_match = RE_HEJI_AMOUNT.search(text)
            if heji_match:
                info['金额'] = heji_match.group(1).replace(',', '')
            else:
                # 第三选择：找所有¥后的金额，取最大的（价税合计通常是最大的）
                # 过滤掉明显异常的小额（<1元）和超大额（>1000万）
                all_amounts = RE_ALL_AMOUNTS.findall(text)
                if all_amounts:
                    amounts_float = []
                    for a in all_amounts:
                        try:
                            amt = float(a.replace(',', ''))
                            if 1.0 < amt < 10000000:
                                amounts_float.append((amt, a))
                        except:
                            pass
                    if amounts_float:
                        max_amount = max(amounts_float, key=lambda x: x[0])
                        info['金额'] = max_amount[1].replace(',', '')

        # 从文件名提取金额（备用方案）
        if not info['金额']:
            filename = os.path.basename(pdf_path)
            amount_match = RE_FILENAME_AMOUNT.search(filename)
            if amount_match:
                info['金额'] = amount_match.group(1)

        # 清理字段中的前缀和后缀
        for key in ['购买方', '销售方']:
            if info.get(key):
                # 去除常见前缀
                value = info[key]
                for prefix in ['名称:', '名称:', '名　　称:', '购买方:', '销售方:', '名　称:', '名 称:']:
                    if value.startswith(prefix):
                        value = value[len(prefix):].strip()
                        break
                # 去除统一社会信用代码等后缀
                for suffix in ['统一社会信用代码/纳税人识别号：', '统一社会信用代码:', '统一社会信用代码/纳税人识别号:']:
                    if suffix in value:
                        value = value.split(suffix)[0].strip()
                        break
                info[key] = value

        return info

    except Exception as e:
        return {
            '发票号码': '', '发票代码': '', '开票日期': '',
            '购买方': '', '购买方税号': '', '销售方': '', '销售方税号': '',
            '项目内容': '', '金额': '',
            '备注': f'解析错误: {str(e)}'
        }


def process_invoices(base_path, buyer_keyword=None, output_path=None, log_callback=None):
    """处理所有发票文件并生成Excel

    Args:
        base_path: 发票文件所在目录
        buyer_keyword: 购买方公司名称关键词（用于识别购买方）
        output_path: 输出Excel文件路径
        log_callback: 日志回调函数，用于GUI模式显示日志
    """
    # 记录开始时间
    start_time = time.time()
    all_invoices = []

    # 定义日志输出函数
    def log(msg):
        if log_callback:
            log_callback(msg)
        else:
            print(msg)

    for root, dirs, files in os.walk(base_path):
        files = [f for f in files if not f.startswith('.')]
        dirs[:] = [d for d in dirs if not d.startswith('.')]

        for file in files:
            # 只处理 PDF 文件（图片不支持 OCR 识别，跳过以避免输出空行）
            if file.lower().endswith('.pdf'):
                full_path = os.path.join(root, file)
                rel_path = os.path.relpath(root, base_path)

                invoice_data = {
                    '文件夹': rel_path,
                    '文件名': file,
                    '文件类型': 'PDF'
                }

                info = extract_invoice_info(full_path, buyer_keyword)
                invoice_data.update(info)

                all_invoices.append(invoice_data)

    all_invoices.sort(key=lambda x: (x['文件夹'], x['文件名']))

    # 检测重复的发票号码，只保留第一次出现的记录
    invoice_tracker = {}
    duplicates = []  # 记录重复的发票信息用于日志
    filtered_invoices = []

    for inv in all_invoices:
        inv_num = inv.get('发票号码', '')
        inv_code = inv.get('发票代码', '')
        # 使用发票号码+发票代码作为唯一标识（如果两者都有）
        unique_key = inv_num
        if inv_num and inv_code:
            unique_key = f"{inv_code}_{inv_num}"

        if unique_key and unique_key in invoice_tracker:
            # 记录重复信息
            original_file = invoice_tracker[unique_key]['文件名']
            duplicate_file = inv['文件名']
            duplicates.append({
                'key': unique_key,
                'original': original_file,
                'duplicate': duplicate_file,
                'inv_num': inv_num,
                'inv_code': inv_code
            })
        else:
            if unique_key:
                invoice_tracker[unique_key] = inv
            filtered_invoices.append(inv)

    # 用过滤后的列表替换原列表
    all_invoices = filtered_invoices

    # 生成Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "发票清单"

    headers = ['序号', '文件夹', '文件名', '发票代码', '发票号码', '开票日期', '购买方', '购买方税号',
               '销售方', '销售方税号', '项目内容', '金额', '备注']

    column_widths = [6, 26, 36, 14, 18, 11, 22, 16, 28, 16, 18, 10, 12]
    for i, w in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64+i)].width = w

    header_font = Font(bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    for idx, inv in enumerate(all_invoices, 1):
        row = idx + 1
        ws.cell(row=row, column=1).value = idx
        ws.cell(row=row, column=2).value = inv['文件夹']
        ws.cell(row=row, column=3).value = inv['文件名']
        ws.cell(row=row, column=4).value = inv.get('发票代码', '')
        ws.cell(row=row, column=5).value = inv.get('发票号码', '')
        ws.cell(row=row, column=6).value = inv.get('开票日期', '')
        ws.cell(row=row, column=7).value = inv.get('购买方', '')
        ws.cell(row=row, column=8).value = inv.get('购买方税号', '')
        ws.cell(row=row, column=9).value = inv.get('销售方', '')
        ws.cell(row=row, column=10).value = inv.get('销售方税号', '')
        ws.cell(row=row, column=11).value = inv.get('项目内容', '')
        ws.cell(row=row, column=12).value = inv.get('金额', '')
        ws.cell(row=row, column=13).value = inv.get('备注', '')

        for col in range(1, 14):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')

    ws.freeze_panes = 'A2'

    if output_path is None:
        output_path = os.path.join(base_path, '发票清单.xlsx')

    wb.save(output_path)

    # 计算耗时
    end_time = time.time()
    total_time = end_time - start_time
    
    # 统计识别率
    pdf_count = sum(1 for inv in all_invoices if inv['文件名'].endswith('.pdf'))
    with_seller = sum(1 for inv in all_invoices if inv.get('销售方') and not inv['销售方'].startswith('*'))
    with_amount = sum(1 for inv in all_invoices if inv.get('金额'))
    with_inv_num = sum(1 for inv in all_invoices if inv.get('发票号码'))

    # 统计总金额
    total_amount = 0
    for inv in all_invoices:
        if inv.get('金额'):
            try:
                total_amount += float(str(inv['金额']).replace(',', ''))
            except:
                pass

    # 按项目内容统计
    item_stats = {}
    for inv in all_invoices:
        item = inv.get('项目内容', '未知')
        if not item:
            item = '未识别'
        amount = inv.get('金额', '0')
        try:
            amt = float(str(amount).replace(',', ''))
        except:
            amt = 0
        if item not in item_stats:
            item_stats[item] = {'count': 0, 'amount': 0}
        item_stats[item]['count'] += 1
        item_stats[item]['amount'] += amt

    # 按销售方统计（Top 10）
    seller_stats = {}
    for inv in all_invoices:
        seller = inv.get('销售方', '未知')
        if not seller or seller.startswith('*'):
            seller = '未识别'
        amount = inv.get('金额', '0')
        try:
            amt = float(str(amount).replace(',', ''))
        except:
            amt = 0
        if seller not in seller_stats:
            seller_stats[seller] = {'count': 0, 'amount': 0}
        seller_stats[seller]['count'] += 1
        seller_stats[seller]['amount'] += amt

    log("=" * 80)
    log(f"{'发票识别完成':^76}")
    log("=" * 80)

    log(f"\n📊 文件统计:")
    log(f"  总文件数: {len(all_invoices)}")
    log(f"  PDF发票数: {pdf_count}")

    log(f"\n📈 识别率统计:")
    if pdf_count > 0:
        log(f"  发票号码识别: {with_inv_num}/{pdf_count} ({with_inv_num/pdf_count*100:.1f}%)")
        log(f"  销售方识别:   {with_seller}/{pdf_count} ({with_seller/pdf_count*100:.1f}%)")
    else:
        log(f"  发票号码识别: N/A (无PDF文件)")
        log(f"  销售方识别:   N/A (无PDF文件)")
    total_count = len(all_invoices)
    if total_count > 0:
        log(f"  金额识别:     {with_amount}/{total_count} ({with_amount/total_count*100:.1f}%)")
    else:
        log(f"  金额识别:     N/A (无文件)")

    log(f"\n💰 金额统计:")
    log(f"  总金额: ¥{total_amount:,.2f}")
    if len(all_invoices) > 0:
        avg_amount = total_amount / len(all_invoices)
        log(f"  平均金额: ¥{avg_amount:,.2f}")

    # 输出重复发票号码的警告
    if duplicates:
        log(f"\n⚠️  重复发票警告: 发现 {len(duplicates)} 条重复发票，已自动去重！")
        log(f"  {'唯一标识':<25} {'原始文件':<30} {'重复文件':<30}")
        log("  " + "-" * 88)
        for d in duplicates[:10]:  # 最多显示10条
            orig_short = d['original'][:27] + '...' if len(d['original']) > 27 else d['original']
            dup_short = d['duplicate'][:27] + '...' if len(d['duplicate']) > 27 else d['duplicate']
            log(f"  {d['key']:<25} {orig_short:<30} {dup_short:<30}")
        if len(duplicates) > 10:
            log(f"  ... (还有 {len(duplicates)-10} 条重复记录未显示)")

    # 输出项目内容统计（Top 10）
    if item_stats:
        log(f"\n📋 项目内容统计 (Top 10):")
        sorted_items = sorted(item_stats.items(), key=lambda x: x[1]['amount'], reverse=True)[:10]
        for item, data in sorted_items:
            item_short = item[:20] + '...' if len(item) > 20 else item
            log(f"  {item_short:<25} {data['count']:>3}张  ¥{data['amount']:>10,.2f}")

    # 输出销售方统计（Top 5）
    if seller_stats:
        log(f"\n🏪 销售方统计 (Top 5):")
        sorted_sellers = sorted(seller_stats.items(), key=lambda x: x[1]['amount'], reverse=True)[:5]
        for seller, data in sorted_sellers:
            seller_short = seller[:20] + '...' if len(seller) > 20 else seller
            log(f"  {seller_short:<25} {data['count']:>3}张  ¥{data['amount']:>10,.2f}")

    log(f"\n⏱️  时间统计:")
    log(f"  总耗时: {total_time:.2f}秒")
    if len(all_invoices) > 0:
        avg_time = total_time / len(all_invoices)
        log(f"  平均每份: {avg_time:.3f}秒")

    log(f"\n💾 Excel已保存: {output_path}")
    log("=" * 80)

    return output_path


if __name__ == "__main__":
    import sys

    # 获取发票目录
    if len(sys.argv) > 1:
        BASE_DIR = sys.argv[1]
    else:
        BASE_DIR = input("请输入发票文件所在目录路径: ").strip()

    # 展开 ~ 目录并清理可能的 shell 转义
    # 注意：只在Unix-like系统上处理\ 转义，避免影响Windows网络路径
    BASE_DIR = os.path.expanduser(BASE_DIR)
    if os.name != 'nt':  # 非Windows系统
        BASE_DIR = BASE_DIR.replace('\\ ', ' ')

    # 获取购买方公司名称关键词
    if len(sys.argv) > 2:
        BUYER_KEYWORD = sys.argv[2]
    else:
        BUYER_KEYWORD = input("请输入购买方公司名称关键词（用于识别购买方）: ").strip()

    # 设置输出路径
    if len(sys.argv) > 3:
        OUTPUT_FILE = sys.argv[3]
    else:
        OUTPUT_FILE = os.path.join(BASE_DIR, '发票清单.xlsx')

    print(f"\n开始处理...")
    print(f"  发票目录: {BASE_DIR}")
    print(f"  购买方关键词: {BUYER_KEYWORD}")
    print(f"  输出文件: {OUTPUT_FILE}\n")

    process_invoices(BASE_DIR, BUYER_KEYWORD, OUTPUT_FILE)
